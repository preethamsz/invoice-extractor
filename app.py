from flask import Flask, request, jsonify, render_template, send_file, send_from_directory
import base64
import json
import re
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ══════════════════════════════════════════════════════════════
# PROMPT
# ══════════════════════════════════════════════════════════════
INVOICE_PROMPT = """You are an expert invoice data extraction AI. Analyze this invoice image and extract ALL key attributes.

Return ONLY a valid JSON object with these fields (use null if not found):
{
  "invoice_number": "",
  "invoice_date": "",
  "due_date": "",
  "vendor": {
    "name": "",
    "address": "",
    "email": "",
    "phone": "",
    "tax_id": ""
  },
  "bill_to": {
    "name": "",
    "address": "",
    "email": "",
    "phone": ""
  },
  "line_items": [
    {
      "description": "",
      "quantity": "",
      "unit_price": "",
      "total": ""
    }
  ],
  "subtotal": "",
  "tax_rate": "",
  "tax_amount": "",
  "discount": "",
  "shipping": "",
  "total_amount": "",
  "currency": "",
  "payment_terms": "",
  "payment_method": "",
  "bank_details": "",
  "notes": "",
  "po_number": ""
}
Return ONLY the JSON, no markdown, no explanation."""


# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════
def extract_json(text):
    match = re.search(r'\{.*\}', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    return None

def clean_json(text):
    text = text.strip()
    text = re.sub(r'^```(?:json)?\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    return text.strip()

def convert_pdf_to_images(pdf_bytes):
    try:
        import fitz
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        result = []
        for page in doc:
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)
            b64 = base64.b64encode(pix.tobytes("png")).decode('utf-8')
            result.append(b64)
        doc.close()
        return result, None
    except ImportError:
        return None, "PyMuPDF not installed. Run: pip install pymupdf"
    except Exception as e:
        return None, str(e)


# ══════════════════════════════════════════════════════════════
# AI PROVIDERS
# ══════════════════════════════════════════════════════════════

def call_groq(api_key, image_b64, mime_type):
    from groq import Groq
    client = Groq(api_key=api_key)
    response = client.chat.completions.create(
        model="meta-llama/llama-4-scout-17b-16e-instruct",
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url",
                 "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}},
                {"type": "text", "text": INVOICE_PROMPT}
            ]
        }],
        max_tokens=2000,
        temperature=0.1
    )
    return response.choices[0].message.content


def call_openai(api_key, image_b64, mime_type):
    from openai import OpenAI
    client = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url",
                 "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}},
                {"type": "text", "text": INVOICE_PROMPT}
            ]
        }],
        max_tokens=2000
    )
    return response.choices[0].message.content


def call_gemini(api_key, image_b64, mime_type):
    import google.generativeai as genai
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel('gemini-2.0-flash')
    image_part = {"mime_type": mime_type, "data": image_b64}
    response = model.generate_content([INVOICE_PROMPT, image_part])
    return response.text


def call_anthropic(api_key, image_b64, mime_type):
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=2000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": mime_type,
                        "data": image_b64
                    }
                },
                {"type": "text", "text": INVOICE_PROMPT}
            ]
        }]
    )
    return response.content[0].text


def call_nvidia(api_key, image_b64, mime_type):
    from openai import OpenAI
    client = OpenAI(
        base_url="https://integrate.api.nvidia.com/v1",
        api_key=api_key
    )
    response = client.chat.completions.create(
        model="meta/llama-4-scout-17b-16e-instruct",
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url",
                 "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}},
                {"type": "text", "text": INVOICE_PROMPT}
            ]
        }],
        max_tokens=2000,
        temperature=0.1
    )
    return response.choices[0].message.content


def call_deepseek(api_key, image_b64, mime_type):
    from openai import OpenAI
    client = OpenAI(
        base_url="https://api.deepseek.com/v1",
        api_key=api_key
    )
    # DeepSeek uses text description for images
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[{
            "role": "user",
            "content": f"I have an invoice image encoded in base64. Please analyze it and extract the data.\n\nImage data (base64): data:{mime_type};base64,{image_b64[:100]}...\n\n{INVOICE_PROMPT}"
        }],
        max_tokens=2000,
        temperature=0.1
    )
    return response.choices[0].message.content


def call_mistral(api_key, image_b64, mime_type):
    from mistralai import Mistral
    client = Mistral(api_key=api_key)
    response = client.chat.complete(
        model="pixtral-large-latest",
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url",
                 "image_url": f"data:{mime_type};base64,{image_b64}"},
                {"type": "text", "text": INVOICE_PROMPT}
            ]
        }]
    )
    return response.choices[0].message.content


def call_cohere(api_key, image_b64, mime_type):
    import cohere
    client = cohere.Client(api_key=api_key)
    response = client.chat(
        model="command-r-plus",
        message=f"Analyze this invoice image (base64 encoded) and extract data.\nImage: data:{mime_type};base64,{image_b64[:200]}...\n\n{INVOICE_PROMPT}"
    )
    return response.text


# Provider router
PROVIDERS = {
    "groq":      call_groq,
    "openai":    call_openai,
    "gemini":    call_gemini,
    "anthropic": call_anthropic,
    "nvidia":    call_nvidia,
    "deepseek":  call_deepseek,
    "mistral":   call_mistral,
    "cohere":    call_cohere,
}


# ══════════════════════════════════════════════════════════════
# PWA ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

@app.route('/sw.js')
def service_worker():
    return send_from_directory('static', 'sw.js',
                               mimetype='application/javascript')

@app.route('/manifest.json')
def manifest():
    return send_from_directory('static', 'manifest.json',
                               mimetype='application/manifest+json')


# ══════════════════════════════════════════════════════════════
# MAIN ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/convert-pdf', methods=['POST'])
def convert_pdf():
    file = request.files.get('pdf')
    if not file:
        return jsonify({'error': 'No PDF uploaded'}), 400
    images, error = convert_pdf_to_images(file.read())
    if error:
        return jsonify({'error': error}), 500
    return jsonify({'success': True, 'images': images, 'page_count': len(images)})


@app.route('/extract', methods=['POST'])
def extract_invoice():
    api_key   = request.form.get('api_key', '').strip()
    provider  = request.form.get('provider', 'groq').strip().lower()
    image_b64 = request.form.get('image_b64', '').strip()
    mime_type = request.form.get('mime_type', 'image/png').strip()

    if not api_key:
        return jsonify({'error': 'API key is required.'}), 400

    if image_b64:
        final_b64  = image_b64
        final_mime = mime_type
    else:
        file = request.files.get('invoice')
        if not file:
            return jsonify({'error': 'No invoice uploaded.'}), 400
        supported = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp']
        if file.content_type not in supported:
            return jsonify({'error': 'Please upload JPG, PNG, or WEBP.'}), 400
        final_b64  = base64.b64encode(file.read()).decode('utf-8')
        final_mime = file.content_type

    if provider not in PROVIDERS:
        return jsonify({'error': f'Unknown provider: {provider}'}), 400

    try:
        raw_text    = PROVIDERS[provider](api_key, final_b64, final_mime)
        result_text = clean_json(raw_text)

        try:
            data = json.loads(result_text)
        except json.JSONDecodeError:
            data = extract_json(result_text)
            if not data:
                return jsonify({'error': 'Could not parse invoice data. Try again.'}), 500

        return jsonify({'success': True, 'data': data, 'provider': provider})

    except ImportError as e:
        pkg = str(e).split("'")[1] if "'" in str(e) else str(e)
        return jsonify({'error': f'Package not installed. Run: pip install {pkg}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download-excel', methods=['POST'])
def download_excel():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        buf = build_excel(data)
        invoice_num = data.get('invoice_number', 'invoice') or 'invoice'
        filename = f"invoice_{re.sub(r'[^a-zA-Z0-9]', '_', str(invoice_num))}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════
def build_excel(data):
    wb  = Workbook()
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def hf(c='1a1a2e'): return PatternFill('solid', start_color=c)
    def fnt(bold=False, color='333333', size=10):
        return Font(name='Arial', bold=bold, color=color, size=size)

    ws = wb.active
    ws.title = 'Invoice Summary'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 44

    r = 1
    ws.merge_cells(f'A{r}:B{r}')
    c = ws.cell(r, 1, 'INVOICE EXTRACTION REPORT')
    c.font = fnt(True, 'FFFFFF', 12)
    c.fill = hf('1a1a2e')
    c.alignment = center
    ws.row_dimensions[r].height = 32

    def write_section(title, fields, start):
        rr = start
        ws.merge_cells(f'A{rr}:B{rr}')
        c2 = ws.cell(rr, 1, title)
        c2.font = fnt(True, 'FFFFFF', 10)
        c2.fill = hf('e63946')
        c2.alignment = center
        ws.row_dimensions[rr].height = 22
        rr += 1
        for idx, (lbl, val) in enumerate(fields):
            bg = 'F8F8F8' if idx % 2 == 0 else 'FFFFFF'
            lc = ws.cell(rr, 1, lbl)
            lc.font = fnt(True, '1a1a2e')
            lc.fill = hf(bg)
            lc.alignment = left
            lc.border = border
            vc = ws.cell(rr, 2, str(val) if val else '—')
            vc.font = fnt(color='333333')
            vc.fill = hf(bg)
            vc.alignment = left
            vc.border = border
            ws.row_dimensions[rr].height = 20
            rr += 1
        return rr + 1

    v = data.get('vendor') or {}
    b = data.get('bill_to') or {}

    r = write_section('INVOICE DETAILS', [
        ('Invoice Number', data.get('invoice_number')),
        ('Invoice Date',   data.get('invoice_date')),
        ('Due Date',       data.get('due_date')),
        ('PO Number',      data.get('po_number')),
        ('Currency',       data.get('currency')),
        ('Payment Terms',  data.get('payment_terms')),
    ], r + 2)

    r = write_section('VENDOR', [
        ('Name',    v.get('name')),
        ('Address', v.get('address')),
        ('Email',   v.get('email')),
        ('Phone',   v.get('phone')),
        ('Tax ID',  v.get('tax_id')),
    ], r)

    r = write_section('BILLED TO', [
        ('Name',    b.get('name')),
        ('Address', b.get('address')),
        ('Email',   b.get('email')),
        ('Phone',   b.get('phone')),
    ], r)

    r = write_section('FINANCIALS', [
        ('Subtotal',       data.get('subtotal')),
        ('Tax Rate',       data.get('tax_rate')),
        ('Tax Amount',     data.get('tax_amount')),
        ('Discount',       data.get('discount')),
        ('Shipping',       data.get('shipping')),
        ('Payment Method', data.get('payment_method')),
        ('Bank Details',   data.get('bank_details')),
        ('Notes',          data.get('notes')),
    ], r)

    ws.merge_cells(f'A{r}:A{r}')
    gtl = ws.cell(r, 1, 'TOTAL AMOUNT')
    gtl.font = fnt(True, '856404', 11)
    gtl.fill = hf('fff3cd')
    gtl.alignment = center
    gtl.border = border
    gtv = ws.cell(r, 2, str(data.get('total_amount') or '—'))
    gtv.font = fnt(True, 'e63946', 14)
    gtv.fill = hf('fff3cd')
    gtv.alignment = center
    gtv.border = border
    ws.row_dimensions[r].height = 28

    ws2 = wb.create_sheet('Line Items')
    ws2.sheet_view.showGridLines = False
    for i, w in enumerate([6, 40, 12, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.merge_cells('A1:E1')
    t = ws2.cell(1, 1, 'LINE ITEMS — TABULAR DATA')
    t.font = fnt(True, 'FFFFFF', 12)
    t.fill = hf('1a1a2e')
    t.alignment = center
    ws2.row_dimensions[1].height = 30

    for ci, h in enumerate(['#', 'Description', 'Quantity', 'Unit Price', 'Total'], 1):
        c3 = ws2.cell(2, ci, h)
        c3.font = fnt(True, 'FFFFFF', 10)
        c3.fill = hf('e63946')
        c3.alignment = center
        c3.border = border
    ws2.row_dimensions[2].height = 22

    items = data.get('line_items') or []
    for ri, item in enumerate(items):
        row_r = ri + 3
        bg = 'F8F8F8' if ri % 2 == 0 else 'FFFFFF'
        vals = [ri+1, item.get('description'),
                item.get('quantity'), item.get('unit_price'), item.get('total')]
        for ci, val in enumerate(vals, 1):
            c4 = ws2.cell(row_r, ci, str(val) if val else '—')
            c4.font = fnt(color='333333')
            c4.fill = hf(bg)
            c4.alignment = center if ci != 2 else left
            c4.border = border
        ws2.row_dimensions[row_r].height = 20

    last = len(items) + 3
    for oi, (lbl, val) in enumerate([
        ('Subtotal', data.get('subtotal')),
        ('Tax',      data.get('tax_amount')),
        ('Discount', data.get('discount')),
        ('Shipping', data.get('shipping'))
    ]):
        rr = last + oi
        ws2.merge_cells(f'A{rr}:C{rr}')
        ws2.cell(rr, 1, '').border = border
        lc = ws2.cell(rr, 4, lbl)
        lc.font = fnt(True, '1a1a2e')
        lc.fill = hf('F8F8F8')
        lc.alignment = center
        lc.border = border
        vc = ws2.cell(rr, 5, str(val) if val else '—')
        vc.font = fnt(color='333333')
        vc.fill = hf('F8F8F8')
        vc.alignment = center
        vc.border = border
        ws2.row_dimensions[rr].height = 20

    gr = last + 4
    ws2.merge_cells(f'A{gr}:C{gr}')
    ws2.cell(gr, 1, '').fill = hf('fff3cd')
    ws2.cell(gr, 1).border = border
    gl = ws2.cell(gr, 4, 'TOTAL AMOUNT')
    gl.font = fnt(True, '856404', 11)
    gl.fill = hf('fff3cd')
    gl.alignment = center
    gl.border = border
    gv = ws2.cell(gr, 5, str(data.get('total_amount') or '—'))
    gv.font = fnt(True, 'e63946', 13)
    gv.fill = hf('fff3cd')
    gv.alignment = center
    gv.border = border
    ws2.row_dimensions[gr].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


if __name__ == '__main__':
    import threading
    import webbrowser
    import sys
    import os

    # Fix path for PyInstaller exe
    if getattr(sys, 'frozen', False):
        # Running as compiled .exe
        base_path = os.path.dirname(sys.executable)
    else:
        # Running as normal python script
        base_path = os.path.dirname(os.path.abspath(__file__))

    # Set template and static folders relative to exe location
    app.template_folder = os.path.join(base_path, 'templates')
    app.static_folder   = os.path.join(base_path, 'static')

    # Auto open browser after 1.5 seconds
    def open_browser():
        import time
        time.sleep(1.5)
        webbrowser.open('http://127.0.0.1:5000')

    threading.Thread(target=open_browser, daemon=True).start()

    # Run Flask
    app.run(debug=False, host='127.0.0.1', port=5000)