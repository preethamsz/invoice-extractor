"""
Invoice Extractor Desktop App
Professional Tkinter UI with Groq AI Backend
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import json
import base64
import re
import os
import io
from datetime import datetime

# ── Third party ────────────────────────────────────────────
try:
    from groq import Groq
except ImportError:
    Groq = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False


# ══════════════════════════════════════════════════════════════
# THEME
# ══════════════════════════════════════════════════════════════
BG        = "#080810"
SURFACE   = "#0f0f1a"
SURFACE2  = "#161625"
SURFACE3  = "#1e1e30"
BORDER    = "#252538"
ACCENT    = "#e63946"
ACCENT2   = "#ff8c69"
ACCENT3   = "#4ecb8d"
TEXT      = "#eeeef5"
MUTED     = "#7070a0"
WHITE     = "#ffffff"

FONT_TITLE  = ("Segoe UI", 22, "bold")
FONT_HEAD   = ("Segoe UI", 13, "bold")
FONT_BODY   = ("Segoe UI", 10)
FONT_SMALL  = ("Segoe UI", 9)
FONT_MONO   = ("Consolas", 10)
FONT_MONO_S = ("Consolas", 9)
FONT_BIG    = ("Segoe UI", 28, "bold")


# ══════════════════════════════════════════════════════════════
# CUSTOM WIDGETS
# ══════════════════════════════════════════════════════════════

class RoundedFrame(tk.Canvas):
    """A rounded rectangle frame widget."""
    def __init__(self, parent, radius=12, bg=SURFACE, border_color=BORDER,
                 border_width=1, **kwargs):
        width  = kwargs.pop('width', 400)
        height = kwargs.pop('height', 100)
        super().__init__(parent, width=width, height=height,
                         bg=parent.cget('bg'), highlightthickness=0, **kwargs)
        self.radius       = radius
        self.fill_color   = bg
        self.border_color = border_color
        self.border_width = border_width
        self._draw(width, height)
        self.bind("<Configure>", self._on_resize)

    def _draw(self, w, h):
        self.delete("all")
        r = self.radius
        bw = self.border_width
        # Border
        self._rounded_rect(bw, bw, w-bw, h-bw, r, self.border_color)
        # Fill
        self._rounded_rect(bw+1, bw+1, w-bw-1, h-bw-1, r-1, self.fill_color)

    def _rounded_rect(self, x1, y1, x2, y2, r, color):
        self.create_arc(x1, y1, x1+2*r, y1+2*r, start=90,  extent=90,  fill=color, outline=color)
        self.create_arc(x2-2*r, y1, x2, y1+2*r, start=0,   extent=90,  fill=color, outline=color)
        self.create_arc(x1, y2-2*r, x1+2*r, y2, start=180, extent=90,  fill=color, outline=color)
        self.create_arc(x2-2*r, y2-2*r, x2, y2, start=270, extent=90,  fill=color, outline=color)
        self.create_rectangle(x1+r, y1, x2-r, y2, fill=color, outline=color)
        self.create_rectangle(x1, y1+r, x2, y2-r, fill=color, outline=color)

    def _on_resize(self, e):
        self._draw(e.width, e.height)


class StyledButton(tk.Label):
    """Flat styled button with hover effect."""
    def __init__(self, parent, text, command=None, style="primary",
                 width=None, font=None, **kwargs):
        colors = {
            "primary": (ACCENT,   WHITE,   "#ff4d5a"),
            "success": (SURFACE2, ACCENT3, SURFACE3),
            "ghost":   (SURFACE2, TEXT,    SURFACE3),
            "danger":  (SURFACE2, ACCENT,  SURFACE3),
        }
        self.bg_normal, self.fg, self.bg_hover = colors.get(style, colors["primary"])
        self.command = command

        cfg = dict(
            text=text,
            bg=self.bg_normal,
            fg=self.fg,
            font=font or ("Segoe UI", 10, "bold"),
            cursor="hand2",
            padx=18, pady=10,
            relief="flat",
            bd=0,
        )
        if width:
            cfg["width"] = width
        cfg.update(kwargs)
        super().__init__(parent, **cfg)

        self.bind("<Enter>",    self._hover_on)
        self.bind("<Leave>",    self._hover_off)
        self.bind("<Button-1>", self._click)

    def _hover_on(self, _):  self.config(bg=self.bg_hover)
    def _hover_off(self, _): self.config(bg=self.bg_normal)

    def _click(self, _):
        if self.command:
            self.command()

    def set_state(self, enabled):
        if enabled:
            self.config(cursor="hand2", fg=self.fg)
            self.bind("<Button-1>", self._click)
        else:
            self.config(cursor="", fg=MUTED)
            self.unbind("<Button-1>")


class SectionCard(tk.Frame):
    """A titled section card."""
    def __init__(self, parent, title, dot_color=ACCENT, **kwargs):
        super().__init__(parent, bg=SURFACE, bd=0, **kwargs)

        # Header bar
        hdr = tk.Frame(self, bg=SURFACE2)
        hdr.pack(fill="x")

        dot = tk.Label(hdr, text="●", fg=dot_color, bg=SURFACE2,
                       font=("Segoe UI", 8))
        dot.pack(side="left", padx=(16, 6), pady=10)

        tk.Label(hdr, text=title.upper(),
                 fg=MUTED, bg=SURFACE2,
                 font=("Consolas", 9)).pack(side="left", pady=10)

        # Separator
        sep = tk.Frame(self, bg=BORDER, height=1)
        sep.pack(fill="x")

        # Content area
        self.content = tk.Frame(self, bg=SURFACE)
        self.content.pack(fill="both", expand=True)


class FieldRow(tk.Frame):
    """A label + value row for non-tabular display."""
    def __init__(self, parent, label, value, alt=False, **kwargs):
        bg = SURFACE3 if alt else SURFACE
        super().__init__(parent, bg=bg, **kwargs)

        tk.Label(self, text=label, fg=MUTED, bg=bg,
                 font=FONT_MONO_S, width=22, anchor="w",
                 padx=16, pady=8).pack(side="left")

        tk.Frame(self, bg=BORDER, width=1).pack(side="left", fill="y", pady=4)

        val = str(value) if value not in [None, "", "null"] else "—"
        tk.Label(self, text=val, fg=TEXT, bg=bg,
                 font=FONT_BODY, anchor="w",
                 padx=16, pady=8, wraplength=420).pack(side="left", fill="x", expand=True)

        tk.Frame(self, bg=BORDER, height=1).pack(side="bottom", fill="x")


# ══════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════

class InvoiceExtractorApp:
    def __init__(self, root):
        self.root        = root
        self.api_key     = tk.StringVar()
        self.file_path   = None
        self.pdf_images  = []   # list of base64 strings
        self.result_data = None
        self.show_key    = False

        self._setup_window()
        self._build_ui()

    # ── Window setup ────────────────────────────────────────
    def _setup_window(self):
        self.root.title("Invoice Extractor · Groq AI")
        self.root.geometry("940x780")
        self.root.minsize(860, 680)
        self.root.configure(bg=BG)

        # Center on screen
        self.root.update_idletasks()
        w = self.root.winfo_width()
        h = self.root.winfo_height()
        x = (self.root.winfo_screenwidth()  // 2) - (w // 2)
        y = (self.root.winfo_screenheight() // 2) - (h // 2)
        self.root.geometry(f"+{x}+{y}")

        # Style ttk scrollbar
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Vertical.TScrollbar",
                        background=SURFACE2, troughcolor=SURFACE,
                        bordercolor=BORDER, arrowcolor=MUTED,
                        relief="flat")

    # ── Build entire UI ──────────────────────────────────────
    def _build_ui(self):
        # ── Top title bar ──
        title_bar = tk.Frame(self.root, bg=SURFACE, pady=0)
        title_bar.pack(fill="x")

        inner = tk.Frame(title_bar, bg=SURFACE)
        inner.pack(padx=28, pady=16)

        tk.Label(inner, text="⚡", fg=ACCENT, bg=SURFACE,
                 font=("Segoe UI", 18)).pack(side="left", padx=(0,10))

        tk.Label(inner, text="Invoice Extractor",
                 fg=TEXT, bg=SURFACE, font=FONT_TITLE).pack(side="left")

        tk.Label(inner, text="  v2.0 · Groq AI · Llama 4",
                 fg=MUTED, bg=SURFACE, font=FONT_SMALL).pack(side="left", pady=(6,0))

        tk.Frame(title_bar, bg=BORDER, height=1).pack(fill="x")

        # ── Main layout: left panel + right results ──
        main = tk.Frame(self.root, bg=BG)
        main.pack(fill="both", expand=True, padx=0, pady=0)

        # Left panel (fixed width)
        self.left_panel = tk.Frame(main, bg=BG, width=420)
        self.left_panel.pack(side="left", fill="y", padx=0, pady=0)
        self.left_panel.pack_propagate(False)

        # Scrollable left
        left_canvas = tk.Canvas(self.left_panel, bg=BG,
                                highlightthickness=0, width=420)
        left_scroll = ttk.Scrollbar(self.left_panel, orient="vertical",
                                    command=left_canvas.yview)
        left_canvas.configure(yscrollcommand=left_scroll.set)
        left_scroll.pack(side="right", fill="y")
        left_canvas.pack(side="left", fill="both", expand=True)

        self.left_frame = tk.Frame(left_canvas, bg=BG)
        left_win = left_canvas.create_window((0,0), window=self.left_frame,
                                              anchor="nw", width=400)

        def _on_frame_configure(e):
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        self.left_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(e):
            left_canvas.itemconfig(left_win, width=e.width)
        left_canvas.bind("<Configure>", _on_canvas_configure)

        # Mouse wheel
        def _on_mousewheel(e):
            left_canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        left_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Divider
        tk.Frame(main, bg=BORDER, width=1).pack(side="left", fill="y")

        # Right panel (results)
        self.right_panel = tk.Frame(main, bg=BG)
        self.right_panel.pack(side="left", fill="both", expand=True)

        self._build_left_panel()
        self._build_right_panel()

    # ── LEFT PANEL ───────────────────────────────────────────
    def _build_left_panel(self):
        p = self.left_frame

        # ── Step 1: API Key ──
        self._section_title(p, "01", "Authentication")

        key_frame = tk.Frame(p, bg=SURFACE, relief="flat")
        key_frame.pack(fill="x", padx=18, pady=(0,6))

        tk.Label(key_frame, text="🔑  Groq API Key",
                 fg=MUTED, bg=SURFACE, font=FONT_SMALL).pack(anchor="w", padx=14, pady=(12,4))

        row = tk.Frame(key_frame, bg=SURFACE)
        row.pack(fill="x", padx=12, pady=(0,8))

        self.key_entry = tk.Entry(row, textvariable=self.api_key,
                                  show="•", bg=SURFACE2, fg=TEXT,
                                  insertbackground=TEXT,
                                  font=FONT_MONO, relief="flat",
                                  bd=0, highlightthickness=1,
                                  highlightbackground=BORDER,
                                  highlightcolor=ACCENT)
        self.key_entry.pack(side="left", fill="x", expand=True,
                            ipady=8, ipadx=10)

        self.show_btn = StyledButton(row, "Show", command=self._toggle_key,
                                     style="ghost", font=FONT_SMALL)
        self.show_btn.pack(side="left", padx=(6,0))

        tk.Label(key_frame,
                 text="Get free key → console.groq.com/keys",
                 fg=MUTED, bg=SURFACE, font=FONT_SMALL,
                 cursor="hand2").pack(anchor="w", padx=14, pady=(0,10))

        self._add_border(p)

        # ── Step 2: Upload ──
        self._section_title(p, "02", "Upload Invoice")

        upload_frame = tk.Frame(p, bg=SURFACE)
        upload_frame.pack(fill="x", padx=18, pady=(0,6))

        # Tabs row
        tabs = tk.Frame(upload_frame, bg=SURFACE)
        tabs.pack(fill="x", padx=12, pady=(12,10))

        self.tab_var = tk.StringVar(value="pdf")

        self.tab_pdf = StyledButton(tabs, "📄  PDF Invoice",
                                    command=lambda: self._switch_tab("pdf"),
                                    style="primary", font=FONT_SMALL)
        self.tab_pdf.pack(side="left", fill="x", expand=True, padx=(0,4))

        self.tab_img = StyledButton(tabs, "🖼️  Image Invoice",
                                    command=lambda: self._switch_tab("img"),
                                    style="ghost", font=FONT_SMALL)
        self.tab_img.pack(side="left", fill="x", expand=True, padx=(4,0))

        # PDF panel
        self.pdf_panel = tk.Frame(upload_frame, bg=SURFACE)
        self.pdf_panel.pack(fill="x", padx=12)

        self._build_drop_zone(self.pdf_panel, "pdf")

        # Image panel
        self.img_panel = tk.Frame(upload_frame, bg=SURFACE)

        self._build_drop_zone(self.img_panel, "img")

        # File info label
        self.file_label = tk.Label(upload_frame, text="No file selected",
                                   fg=MUTED, bg=SURFACE, font=FONT_SMALL)
        self.file_label.pack(anchor="w", padx=14, pady=(6,10))

        self._add_border(p)

        # ── Step 3: Extract button ──
        self._section_title(p, "03", "Extract")

        btn_frame = tk.Frame(p, bg=BG)
        btn_frame.pack(fill="x", padx=18, pady=(0,10))

        self.extract_btn = tk.Label(
            btn_frame,
            text="⚡   Extract Invoice Data",
            bg=ACCENT, fg=WHITE,
            font=("Segoe UI", 12, "bold"),
            cursor="hand2", pady=14,
            relief="flat"
        )
        self.extract_btn.pack(fill="x")
        self.extract_btn.bind("<Button-1>", lambda e: self._extract())
        self.extract_btn.bind("<Enter>",    lambda e: self.extract_btn.config(bg="#ff4d5a"))
        self.extract_btn.bind("<Leave>",    lambda e: self.extract_btn.config(bg=ACCENT))

        # ── Progress ──
        self.progress_frame = tk.Frame(p, bg=BG)
        self.progress_frame.pack(fill="x", padx=18, pady=(0,10))

        self.progress_steps = []
        steps = [
            ("🔑", "Authenticating with Groq"),
            ("📤", "Sending invoice to AI"),
            ("🧠", "AI analyzing invoice"),
            ("✅", "Processing results"),
        ]
        for icon, label in steps:
            row = tk.Frame(self.progress_frame, bg=SURFACE2, pady=6)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=icon, bg=SURFACE2,
                     font=("Segoe UI", 11)).pack(side="left", padx=(10,6))
            lbl = tk.Label(row, text=label, fg=MUTED,
                           bg=SURFACE2, font=FONT_SMALL)
            lbl.pack(side="left")
            self.progress_steps.append((row, lbl))

        self.progress_frame.pack_forget()

        # ── Download button (hidden until results) ──
        self.download_frame = tk.Frame(p, bg=BG)
        self.download_frame.pack(fill="x", padx=18, pady=(0,16))

        self.download_btn = StyledButton(
            self.download_frame,
            "📥   Download Excel",
            command=self._download_excel,
            style="success",
            font=("Segoe UI", 10, "bold")
        )
        self.download_btn.pack(fill="x", pady=4)

        self.copy_btn = StyledButton(
            self.download_frame,
            "📋   Copy JSON",
            command=self._copy_json,
            style="ghost",
            font=FONT_SMALL
        )
        self.copy_btn.pack(fill="x")
        self.download_frame.pack_forget()

    def _build_drop_zone(self, parent, kind):
        zone = tk.Frame(parent, bg=SURFACE2,
                        highlightthickness=1,
                        highlightbackground=BORDER)
        zone.pack(fill="x", pady=(0,4))

        inner = tk.Frame(zone, bg=SURFACE2)
        inner.pack(pady=24)

        tk.Label(inner, text="⬆", fg=ACCENT, bg=SURFACE2,
                 font=("Segoe UI", 24)).pack()

        hint = "Drop PDF here or click to browse" if kind == "pdf" else "Drop image here (JPG/PNG/WEBP)"
        tk.Label(inner, text=hint, fg=TEXT, bg=SURFACE2,
                 font=("Segoe UI", 10, "bold")).pack(pady=(6,2))

        sub = "PDF auto-converts to images" if kind == "pdf" else "Max 10MB"
        tk.Label(inner, text=sub, fg=MUTED, bg=SURFACE2,
                 font=FONT_SMALL).pack()

        # Click to browse
        cmd = self._browse_pdf if kind == "pdf" else self._browse_img
        for w in [zone, inner]:
            w.bind("<Button-1>", lambda e, c=cmd: c())
            w.configure(cursor="hand2")

        # Hover
        def hover_on(e):  zone.configure(highlightbackground=ACCENT)
        def hover_off(e): zone.configure(highlightbackground=BORDER)
        zone.bind("<Enter>", hover_on)
        zone.bind("<Leave>", hover_off)

    # ── RIGHT PANEL ──────────────────────────────────────────
    def _build_right_panel(self):
        p = self.right_panel

        # Placeholder
        self.placeholder = tk.Frame(p, bg=BG)
        self.placeholder.pack(fill="both", expand=True)

        tk.Label(self.placeholder, text="🧾", fg=SURFACE3,
                 bg=BG, font=("Segoe UI", 64)).pack(expand=True, pady=(120,8))
        tk.Label(self.placeholder, text="Invoice results will appear here",
                 fg=SURFACE3, bg=BG, font=("Segoe UI", 13)).pack()
        tk.Label(self.placeholder, text="Upload an invoice and click Extract",
                 fg=SURFACE3, bg=BG, font=FONT_SMALL).pack(pady=(4,0))

        # Results scroll area
        self.results_outer = tk.Frame(p, bg=BG)

        canvas = tk.Canvas(self.results_outer, bg=BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.results_outer, orient="vertical",
                                  command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self.results_frame = tk.Frame(canvas, bg=BG)
        self._rwin = canvas.create_window((0,0), window=self.results_frame,
                                           anchor="nw")

        def _cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        self.results_frame.bind("<Configure>", _cfg)

        def _resize(e):
            canvas.itemconfig(self._rwin, width=e.width)
        canvas.bind("<Configure>", _resize)

        def _mw(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas.bind("<MouseWheel>", _mw)

    # ── Helpers ──────────────────────────────────────────────
    def _section_title(self, parent, num, title):
        row = tk.Frame(parent, bg=BG)
        row.pack(fill="x", padx=18, pady=(16,6))

        tk.Label(row, text=num, fg=WHITE, bg=ACCENT,
                 font=("Consolas", 9, "bold"),
                 padx=7, pady=3).pack(side="left")

        tk.Label(row, text=f"  {title.upper()}",
                 fg=TEXT, bg=BG,
                 font=("Segoe UI", 11, "bold")).pack(side="left")

        tk.Frame(row, bg=BORDER, height=1).pack(side="left",
                                                 fill="x", expand=True,
                                                 padx=(12,0), pady=6)

    def _add_border(self, parent):
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x",
                                                    padx=18, pady=8)

    def _toggle_key(self):
        self.show_key = not self.show_key
        self.key_entry.config(show="" if self.show_key else "•")
        self.show_btn.config(text="Hide" if self.show_key else "Show")

    def _switch_tab(self, tab):
        self.tab_var.set(tab)
        if tab == "pdf":
            self.tab_pdf.config(bg=ACCENT, fg=WHITE)
            self.tab_img.config(bg=SURFACE2, fg=TEXT)
            self.img_panel.pack_forget()
            self.pdf_panel.pack(fill="x", padx=12)
        else:
            self.tab_img.config(bg=ACCENT, fg=WHITE)
            self.tab_pdf.config(bg=SURFACE2, fg=TEXT)
            self.pdf_panel.pack_forget()
            self.img_panel.pack(fill="x", padx=12)

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="Select Invoice PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if path:
            self.file_path = path
            name = os.path.basename(path)
            self.file_label.config(text=f"📄 {name}", fg=ACCENT3)
            self._convert_pdf(path)

    def _browse_img(self):
        path = filedialog.askopenfilename(
            title="Select Invoice Image",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.webp"),
                       ("All files", "*.*")]
        )
        if path:
            self.file_path = path
            name = os.path.basename(path)
            self.file_label.config(text=f"🖼️ {name}", fg=ACCENT3)
            self.pdf_images = []

    def _convert_pdf(self, path):
        self.file_label.config(text="⏳ Converting PDF to images...", fg=ACCENT2)
        self.root.update()

        def _do():
            try:
                if not HAS_FITZ:
                    self._set_label("❌ PyMuPDF not installed. Run: pip install pymupdf", ACCENT)
                    return

                doc = fitz.open(path)
                images = []
                for page in doc:
                    mat = fitz.Matrix(2.0, 2.0)
                    pix = page.get_pixmap(matrix=mat)
                    b64 = base64.b64encode(pix.tobytes("png")).decode()
                    images.append(b64)
                doc.close()

                self.pdf_images = images
                count = len(images)
                self.root.after(0, lambda: self._set_label(
                    f"✅ {count} page(s) converted — ready to extract!", ACCENT3))
            except Exception as e:
                self.root.after(0, lambda: self._set_label(f"❌ {e}", ACCENT))

        threading.Thread(target=_do, daemon=True).start()

    def _set_label(self, text, color):
        self.file_label.config(text=text, fg=color)

    # ── EXTRACTION ───────────────────────────────────────────
    def _extract(self):
        api_key = self.api_key.get().strip()
        if not api_key:
            messagebox.showwarning("Missing API Key", "Please enter your Groq API key.")
            return

        tab = self.tab_var.get()

        if tab == "pdf":
            if not self.pdf_images:
                messagebox.showwarning("No Images", "Please select a PDF first.\nIf you already selected one, wait for conversion to finish.")
                return
            image_b64 = self.pdf_images[0]
            mime = "image/png"
        else:
            if not self.file_path:
                messagebox.showwarning("No File", "Please select an image file.")
                return
            with open(self.file_path, "rb") as f:
                image_b64 = base64.b64encode(f.read()).decode()
            ext = os.path.splitext(self.file_path)[1].lower()
            mime = {"jpg": "image/jpeg", ".jpeg": "image/jpeg",
                    ".png": "image/png", ".webp": "image/webp"}.get(ext, "image/jpeg")

        # Show progress
        self._show_progress(True)
        self.extract_btn.config(text="⏳  Analyzing...", bg=SURFACE2, fg=MUTED)
        self.extract_btn.unbind("<Button-1>")
        self.download_frame.pack_forget()

        def _do():
            try:
                self._set_prog(1)
                if not Groq:
                    raise Exception("groq not installed. Run: pip install groq")

                client = Groq(api_key=api_key)
                self._set_prog(2)

                prompt = """You are an expert invoice data extraction AI. Analyze this invoice image and extract ALL key attributes.

Return ONLY a valid JSON object with these fields (use null if not found):
{
  "invoice_number": "",
  "invoice_date": "",
  "due_date": "",
  "vendor": {"name":"","address":"","email":"","phone":"","tax_id":""},
  "bill_to": {"name":"","address":"","email":"","phone":""},
  "line_items": [{"description":"","quantity":"","unit_price":"","total":""}],
  "subtotal": "",
  "tax_rate": "",
  "tax_amount": "",
  "discount": "",
  "shipping": "",
  "total_amount": "",
  "currency": "",
  "payment_terms": "",
  "payment_method": "",
  "bank_details": "",
  "notes": "",
  "po_number": ""
}
Return ONLY the JSON, no markdown, no explanation."""

                self._set_prog(3)
                response = client.chat.completions.create(
                    model="meta-llama/llama-4-scout-17b-16e-instruct",
                    messages=[{
                        "role": "user",
                        "content": [
                            {"type": "image_url",
                             "image_url": {"url": f"data:{mime};base64,{image_b64}"}},
                            {"type": "text", "text": prompt}
                        ]
                    }],
                    max_tokens=2000,
                    temperature=0.1
                )

                self._set_prog(4)
                text = response.choices[0].message.content.strip()
                text = re.sub(r'^```(?:json)?\s*', '', text)
                text = re.sub(r'\s*```$', '', text)
                data = json.loads(text)

                self.result_data = data
                self.root.after(0, lambda: self._show_results(data))

            except Exception as e:
                self.root.after(0, lambda: self._on_error(str(e)))

        threading.Thread(target=_do, daemon=True).start()

    def _set_prog(self, step):
        def _do():
            for i, (row, lbl) in enumerate(self.progress_steps):
                n = i + 1
                if n < step:
                    row.config(bg=SURFACE2)
                    lbl.config(fg=ACCENT3, bg=SURFACE2)
                elif n == step:
                    row.config(bg=SURFACE3)
                    lbl.config(fg=TEXT, bg=SURFACE3)
                else:
                    row.config(bg=SURFACE2)
                    lbl.config(fg=MUTED, bg=SURFACE2)
        self.root.after(0, _do)

    def _show_progress(self, show):
        if show:
            self.progress_frame.pack(fill="x", padx=18, pady=(0,10))
        else:
            self.progress_frame.pack_forget()

    def _on_error(self, msg):
        self._show_progress(False)
        self.extract_btn.config(text="⚡   Extract Invoice Data",
                                 bg=ACCENT, fg=WHITE)
        self.extract_btn.bind("<Button-1>", lambda e: self._extract())
        messagebox.showerror("Extraction Error", msg)

    # ── RESULTS RENDERING ────────────────────────────────────
    def _show_results(self, data):
        self._show_progress(False)
        self.extract_btn.config(text="⚡   Extract Invoice Data",
                                 bg=ACCENT, fg=WHITE)
        self.extract_btn.bind("<Button-1>", lambda e: self._extract())

        # Show download buttons
        self.download_frame.pack(fill="x", padx=18, pady=(0,16))

        # Switch to results panel
        self.placeholder.pack_forget()
        self.results_outer.pack(fill="both", expand=True)

        # Clear old results
        for w in self.results_frame.winfo_children():
            w.destroy()

        pad = dict(padx=16, pady=(0,12))

        # ── Title ──
        hdr = tk.Frame(self.results_frame, bg=BG)
        hdr.pack(fill="x", padx=16, pady=(16,12))

        tk.Label(hdr, text="Extracted Data",
                 fg=TEXT, bg=BG,
                 font=("Segoe UI", 18, "bold")).pack(side="left")

        ts = datetime.now().strftime("%H:%M:%S")
        tk.Label(hdr, text=f"  · {ts}",
                 fg=MUTED, bg=BG,
                 font=FONT_SMALL).pack(side="left", pady=(6,0))

        # ── NON-TABULAR sections ──
        self._result_section(
            "📋  Invoice Overview  [NON-TABULAR]", ACCENT,
            [
                ("Invoice Number",  data.get("invoice_number")),
                ("Invoice Date",    data.get("invoice_date")),
                ("Due Date",        data.get("due_date")),
                ("PO Number",       data.get("po_number")),
                ("Currency",        data.get("currency")),
                ("Payment Terms",   data.get("payment_terms")),
                ("Payment Method",  data.get("payment_method")),
            ]
        )

        v = data.get("vendor") or {}
        self._result_section(
            "🏢  Vendor Details  [NON-TABULAR]", ACCENT,
            [
                ("Name",    v.get("name")),
                ("Address", v.get("address")),
                ("Email",   v.get("email")),
                ("Phone",   v.get("phone")),
                ("Tax ID",  v.get("tax_id")),
            ]
        )

        b = data.get("bill_to") or {}
        self._result_section(
            "📬  Billed To  [NON-TABULAR]", ACCENT,
            [
                ("Name",    b.get("name")),
                ("Address", b.get("address")),
                ("Email",   b.get("email")),
                ("Phone",   b.get("phone")),
            ]
        )

        # ── TABULAR: Line Items ──
        items = data.get("line_items") or []
        if items:
            self._result_table(items)

        # ── Totals ──
        self._result_section(
            "💰  Totals & Payment  [NON-TABULAR]", ACCENT3,
            [
                ("Subtotal",      data.get("subtotal")),
                ("Tax Rate",      data.get("tax_rate")),
                ("Tax Amount",    data.get("tax_amount")),
                ("Discount",      data.get("discount")),
                ("Shipping",      data.get("shipping")),
                ("Bank Details",  data.get("bank_details")),
                ("Notes",         data.get("notes")),
                ("TOTAL AMOUNT",  data.get("total_amount")),
            ],
            highlight_last=True
        )

    def _result_section(self, title, dot, fields, highlight_last=False):
        card = SectionCard(self.results_frame, title, dot_color=dot)
        card.pack(fill="x", padx=16, pady=(0,10))

        for i, (label, value) in enumerate(fields):
            is_last   = (i == len(fields) - 1) and highlight_last
            alt       = i % 2 == 0
            row_bg    = SURFACE3 if alt else SURFACE
            if is_last:
                row_bg = "#1a0d0e"

            row = tk.Frame(card.content, bg=row_bg)
            row.pack(fill="x")

            lbl_fg = ACCENT if is_last else MUTED
            val_fg = ACCENT if is_last else TEXT
            lbl_fn = ("Segoe UI", 9, "bold") if is_last else FONT_MONO_S
            val_fn = ("Segoe UI", 13, "bold") if is_last else FONT_BODY

            tk.Label(row, text=label, fg=lbl_fg, bg=row_bg,
                     font=lbl_fn, width=20, anchor="w",
                     padx=14, pady=9).pack(side="left")

            tk.Frame(row, bg=BORDER, width=1).pack(side="left", fill="y", pady=4)

            val = str(value) if value not in [None, "", "null"] else "—"
            tk.Label(row, text=val, fg=val_fg, bg=row_bg,
                     font=val_fn, anchor="w",
                     padx=14, pady=9,
                     wraplength=340).pack(side="left", fill="x", expand=True)

            tk.Frame(card.content, bg=BORDER, height=1).pack(fill="x")

    def _result_table(self, items):
        card = SectionCard(self.results_frame,
                           "📦  Line Items  [TABULAR DATA]",
                           dot_color=ACCENT2)
        card.pack(fill="x", padx=16, pady=(0,10))

        # Table header
        cols = ["#", "Description", "Qty", "Unit Price", "Total"]
        widths = [3, 28, 8, 12, 12]
        hdr = tk.Frame(card.content, bg=SURFACE3)
        hdr.pack(fill="x")

        for col, w in zip(cols, widths):
            tk.Label(hdr, text=col, fg=MUTED, bg=SURFACE3,
                     font=FONT_MONO_S, width=w, anchor="w",
                     padx=12, pady=8).pack(side="left")

        tk.Frame(card.content, bg=BORDER, height=1).pack(fill="x")

        for i, item in enumerate(items):
            row_bg = SURFACE3 if i % 2 == 0 else SURFACE
            row = tk.Frame(card.content, bg=row_bg)
            row.pack(fill="x")

            vals = [
                str(i+1),
                str(item.get("description") or "—"),
                str(item.get("quantity") or "—"),
                str(item.get("unit_price") or "—"),
                str(item.get("total") or "—"),
            ]

            for val, w in zip(vals, widths):
                tk.Label(row, text=val, fg=TEXT, bg=row_bg,
                         font=FONT_BODY, width=w, anchor="w",
                         padx=12, pady=8).pack(side="left")

            tk.Frame(card.content, bg=BORDER, height=1).pack(fill="x")

    # ── DOWNLOAD EXCEL ───────────────────────────────────────
    def _download_excel(self):
        if not self.result_data:
            messagebox.showwarning("No Data", "Extract an invoice first.")
            return

        if not HAS_EXCEL:
            messagebox.showerror("Missing Package",
                                 "openpyxl not installed.\nRun: pip install openpyxl")
            return

        inv_num = self.result_data.get("invoice_number") or "invoice"
        default = f"invoice_{re.sub(r'[^a-zA-Z0-9]','_', str(inv_num))}.xlsx"

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default,
            title="Save Invoice Excel"
        )

        if not path:
            return

        try:
            self._build_excel(path, self.result_data)
            messagebox.showinfo("Downloaded!", f"Excel saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _build_excel(self, path, data):
        wb = Workbook()
        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_a = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        def hf(color="1a1a2e"): return PatternFill("solid", start_color=color)
        def fnt(bold=False, color="333333", size=10):
            return Font(name="Arial", bold=bold, color=color, size=size)

        # ── Sheet 1: Summary ──
        ws = wb.active
        ws.title = "Invoice Summary"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 44

        r = 1
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(r, 1, "INVOICE EXTRACTION REPORT")
        c.font = fnt(True, "FFFFFF", 12)
        c.fill = hf("1a1a2e")
        c.alignment = center
        ws.row_dimensions[r].height = 32

        def write_section(title, fields, start):
            rr = start
            ws.merge_cells(f"A{rr}:B{rr}")
            c2 = ws.cell(rr, 1, title)
            c2.font = fnt(True, "FFFFFF", 10)
            c2.fill = hf("e63946")
            c2.alignment = center
            ws.row_dimensions[rr].height = 22
            rr += 1
            for idx, (lbl, val) in enumerate(fields):
                bg = "F8F8F8" if idx % 2 == 0 else "FFFFFF"
                lc = ws.cell(rr, 1, lbl)
                lc.font = fnt(True, "1a1a2e")
                lc.fill = hf(bg)
                lc.alignment = left_a
                lc.border = border
                vc = ws.cell(rr, 2, str(val) if val else "—")
                vc.font = fnt(color="333333")
                vc.fill = hf(bg)
                vc.alignment = left_a
                vc.border = border
                ws.row_dimensions[rr].height = 20
                rr += 1
            return rr + 1

        v = data.get("vendor") or {}
        b = data.get("bill_to") or {}

        r = write_section("INVOICE DETAILS", [
            ("Invoice Number", data.get("invoice_number")),
            ("Invoice Date",   data.get("invoice_date")),
            ("Due Date",       data.get("due_date")),
            ("PO Number",      data.get("po_number")),
            ("Currency",       data.get("currency")),
            ("Payment Terms",  data.get("payment_terms")),
        ], r + 2)

        r = write_section("VENDOR", [
            ("Name",    v.get("name")),
            ("Address", v.get("address")),
            ("Email",   v.get("email")),
            ("Phone",   v.get("phone")),
            ("Tax ID",  v.get("tax_id")),
        ], r)

        r = write_section("BILLED TO", [
            ("Name",    b.get("name")),
            ("Address", b.get("address")),
            ("Email",   b.get("email")),
            ("Phone",   b.get("phone")),
        ], r)

        r = write_section("FINANCIALS", [
            ("Subtotal",       data.get("subtotal")),
            ("Tax Rate",       data.get("tax_rate")),
            ("Tax Amount",     data.get("tax_amount")),
            ("Discount",       data.get("discount")),
            ("Shipping",       data.get("shipping")),
            ("Payment Method", data.get("payment_method")),
            ("Bank Details",   data.get("bank_details")),
            ("Notes",          data.get("notes")),
        ], r)

        # Grand total
        ws.merge_cells(f"A{r}:A{r}")
        gtl = ws.cell(r, 1, "TOTAL AMOUNT")
        gtl.font = fnt(True, "856404", 11)
        gtl.fill = hf("fff3cd")
        gtl.alignment = center
        gtl.border = border
        gtv = ws.cell(r, 2, str(data.get("total_amount") or "—"))
        gtv.font = fnt(True, "e63946", 14)
        gtv.fill = hf("fff3cd")
        gtv.alignment = center
        gtv.border = border
        ws.row_dimensions[r].height = 28

        # ── Sheet 2: Line Items ──
        ws2 = wb.create_sheet("Line Items")
        ws2.sheet_view.showGridLines = False
        col_w = [6, 40, 12, 14, 14]
        for i, w2 in enumerate(col_w, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w2

        ws2.merge_cells("A1:E1")
        t = ws2.cell(1, 1, "LINE ITEMS — TABULAR DATA")
        t.font = fnt(True, "FFFFFF", 12)
        t.fill = hf("1a1a2e")
        t.alignment = center
        ws2.row_dimensions[1].height = 30

        hdrs = ["#", "Description", "Quantity", "Unit Price", "Total"]
        for ci, h in enumerate(hdrs, 1):
            c3 = ws2.cell(2, ci, h)
            c3.font = fnt(True, "FFFFFF", 10)
            c3.fill = hf("e63946")
            c3.alignment = center
            c3.border = border
        ws2.row_dimensions[2].height = 22

        items = data.get("line_items") or []
        for ri, item in enumerate(items):
            row_r = ri + 3
            bg = "F8F8F8" if ri % 2 == 0 else "FFFFFF"
            vals = [ri+1, item.get("description"),
                    item.get("quantity"), item.get("unit_price"), item.get("total")]
            for ci, val in enumerate(vals, 1):
                c4 = ws2.cell(row_r, ci, str(val) if val else "—")
                c4.font = fnt(color="333333")
                c4.fill = hf(bg)
                c4.alignment = center if ci != 2 else left_a
                c4.border = border
            ws2.row_dimensions[row_r].height = 20

        # Totals rows
        last = len(items) + 3
        tots = [("Subtotal", data.get("subtotal")),
                ("Tax",      data.get("tax_amount")),
                ("Discount", data.get("discount")),
                ("Shipping", data.get("shipping"))]

        for oi, (lbl, val) in enumerate(tots):
            rr = last + oi
            ws2.merge_cells(f"A{rr}:C{rr}")
            ws2.cell(rr, 1, "").border = border
            lc = ws2.cell(rr, 4, lbl)
            lc.font = fnt(True, "1a1a2e")
            lc.fill = hf("F8F8F8")
            lc.alignment = center
            lc.border = border
            vc = ws2.cell(rr, 5, str(val) if val else "—")
            vc.font = fnt(color="333333")
            vc.fill = hf("F8F8F8")
            vc.alignment = center
            vc.border = border
            ws2.row_dimensions[rr].height = 20

        gr = last + len(tots)
        ws2.merge_cells(f"A{gr}:C{gr}")
        ws2.cell(gr, 1, "").fill = hf("fff3cd")
        ws2.cell(gr, 1).border = border
        gl = ws2.cell(gr, 4, "TOTAL AMOUNT")
        gl.font = fnt(True, "856404", 11)
        gl.fill = hf("fff3cd")
        gl.alignment = center
        gl.border = border
        gv = ws2.cell(gr, 5, str(data.get("total_amount") or "—"))
        gv.font = fnt(True, "e63946", 13)
        gv.fill = hf("fff3cd")
        gv.alignment = center
        gv.border = border
        ws2.row_dimensions[gr].height = 28

        wb.save(path)

    # ── COPY JSON ────────────────────────────────────────────
    def _copy_json(self):
        if not self.result_data:
            messagebox.showwarning("No Data", "Extract an invoice first.")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(json.dumps(self.result_data, indent=2))
        messagebox.showinfo("Copied!", "JSON copied to clipboard.")


# ══════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    app  = InvoiceExtractorApp(root)
    root.mainloop()