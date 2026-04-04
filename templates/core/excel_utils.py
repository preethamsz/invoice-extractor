import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel(path, data):
    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_a = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def hf(color="1a1a2e"): return PatternFill("solid", start_color=color)
    def fnt(bold=False, color="333333", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Invoice Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44

    r = 1
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(r, 1, "INVOICE EXTRACTION REPORT")
    c.font = fnt(True, "FFFFFF", 12)
    c.fill = hf("1a1a2e")
    c.alignment = center
    ws.row_dimensions[r].height = 32

    def write_section(title, fields, start):
        rr = start
        ws.merge_cells(f"A{rr}:B{rr}")
        c2 = ws.cell(rr, 1, title)
        c2.font = fnt(True, "FFFFFF", 10)
        c2.fill = hf("e63946")
        c2.alignment = center
        ws.row_dimensions[rr].height = 22
        rr += 1
        for idx, (lbl, val) in enumerate(fields):
            bg = "F8F8F8" if idx % 2 == 0 else "FFFFFF"
            lc = ws.cell(rr, 1, lbl)
            lc.font = fnt(True, "1a1a2e")
            lc.fill = hf(bg)
            lc.alignment = left_a
            lc.border = border
            vc = ws.cell(rr, 2, str(val) if val else "—")
            vc.font = fnt(color="333333")
            vc.fill = hf(bg)
            vc.alignment = left_a
            vc.border = border
            ws.row_dimensions[rr].height = 20
            rr += 1
        return rr + 1

    v = data.get("vendor") or {}
    b = data.get("bill_to") or {}

    r = write_section("INVOICE DETAILS", [
        ("Invoice Number", data.get("invoice_number")),
        ("Invoice Date",   data.get("invoice_date")),
        ("Due Date",       data.get("due_date")),
        ("PO Number",      data.get("po_number")),
        ("Currency",       data.get("currency")),
        ("Payment Terms",  data.get("payment_terms")),
    ], r + 2)

    r = write_section("VENDOR", [
        ("Name",    v.get("name")),
        ("Address", v.get("address")),
        ("Email",   v.get("email")),
        ("Phone",   v.get("phone")),
        ("Tax ID",  v.get("tax_id")),
    ], r)

    r = write_section("BILLED TO", [
        ("Name",    b.get("name")),
        ("Address", b.get("address")),
        ("Email",   b.get("email")),
        ("Phone",   b.get("phone")),
    ], r)

    r = write_section("FINANCIALS", [
        ("Subtotal",       data.get("subtotal")),
        ("Tax Rate",       data.get("tax_rate")),
        ("Tax Amount",     data.get("tax_amount")),
        ("Discount",       data.get("discount")),
        ("Shipping",       data.get("shipping")),
        ("Payment Method", data.get("payment_method")),
        ("Bank Details",   data.get("bank_details")),
        ("Notes",          data.get("notes")),
    ], r)

    # Grand total
    ws.merge_cells(f"A{r}:A{r}")
    gtl = ws.cell(r, 1, "TOTAL AMOUNT")
    gtl.font = fnt(True, "856404", 11)
    gtl.fill = hf("fff3cd")
    gtl.alignment = center
    gtl.border = border
    gtv = ws.cell(r, 2, str(data.get("total_amount") or "—"))
    gtv.font = fnt(True, "e63946", 14)
    gtv.fill = hf("fff3cd")
    gtv.alignment = center
    gtv.border = border
    ws.row_dimensions[r].height = 28

    # ── Sheet 2: Line Items ──
    ws2 = wb.create_sheet("Line Items")
    ws2.sheet_view.showGridLines = False
    col_w = [6, 40, 12, 14, 14]
    for i, w2 in enumerate(col_w, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w2

    ws2.merge_cells("A1:E1")
    t = ws2.cell(1, 1, "LINE ITEMS — TABULAR DATA")
    t.font = fnt(True, "FFFFFF", 12)
    t.fill = hf("1a1a2e")
    t.alignment = center
    ws2.row_dimensions[1].height = 30

    hdrs = ["#", "Description", "Quantity", "Unit Price", "Total"]
    for ci, h in enumerate(hdrs, 1):
        c3 = ws2.cell(2, ci, h)
        c3.font = fnt(True, "FFFFFF", 10)
        c3.fill = hf("e63946")
        c3.alignment = center
        c3.border = border
    ws2.row_dimensions[2].height = 22

    items = data.get("line_items") or []
    for ri, item in enumerate(items):
        row_r = ri + 3
        bg = "F8F8F8" if ri % 2 == 0 else "FFFFFF"
        vals = [ri+1, item.get("description"),
                item.get("quantity"), item.get("unit_price"), item.get("total")]
        for ci, val in enumerate(vals, 1):
            c4 = ws2.cell(row_r, ci, str(val) if val else "—")
            c4.font = fnt(color="333333")
            c4.fill = hf(bg)
            c4.alignment = center if ci != 2 else left_a
            c4.border = border
        ws2.row_dimensions[row_r].height = 20

    # Totals rows
    last = len(items) + 3
    tots = [("Subtotal", data.get("subtotal")),
            ("Tax",      data.get("tax_amount")),
            ("Discount", data.get("discount")),
            ("Shipping", data.get("shipping"))]

    for oi, (lbl, val) in enumerate(tots):
        rr = last + oi
        ws2.merge_cells(f"A{rr}:C{rr}")
        ws2.cell(rr, 1, "").border = border
        lc = ws2.cell(rr, 4, lbl)
        lc.font = fnt(True, "1a1a2e")
        lc.fill = hf("F8F8F8")
        lc.alignment = center
        lc.border = border
        vc = ws2.cell(rr, 5, str(val) if val else "—")
        vc.font = fnt(color="333333")
        vc.fill = hf("F8F8F8")
        vc.alignment = center
        vc.border = border
        ws2.row_dimensions[rr].height = 20

    gr = last + len(tots)
    ws2.merge_cells(f"A{gr}:C{gr}")
    ws2.cell(gr, 1, "").fill = hf("fff3cd")
    ws2.cell(gr, 1).border = border
    gl = ws2.cell(gr, 4, "TOTAL AMOUNT")
    gl.font = fnt(True, "856404", 11)
    gl.fill = hf("fff3cd")
    gl.alignment = center
    gl.border = border
    gv = ws2.cell(gr, 5, str(data.get("total_amount") or "—"))
    gv.font = fnt(True, "e63946", 13)
    gv.fill = hf("fff3cd")
    gv.alignment = center
    gv.border = border
    ws2.row_dimensions[gr].height = 28

    wb.save(path)